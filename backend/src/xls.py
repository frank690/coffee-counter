"""
Handling Excel file operations.
"""

__all__ = ["initialize_spreadsheet", "get_users", "user_buys_coffee"]


import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

EXCEL_FILE = "data.xlsx"


def initialize_spreadsheet():
    """
    Initializes a new Excel file with required sheets if it doesn't exist.
    """
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "users"
        ws_users.append(["uid", "name", "count"])

        ws_actions = wb.create_sheet("actions")
        ws_actions.append(["uid", "name", "date"])

        wb.save(EXCEL_FILE)


def _get_spreadsheet():
    """
    Loads and returns the Excel workbook.

    Returns:
        Workbook: The loaded Excel workbook.
    """
    return load_workbook(EXCEL_FILE)


def get_users() -> list:
    """
    Get all users from the spreadsheet.

    Returns:
        list: A list of user dictionaries.
    """
    wb = _get_spreadsheet()
    ws = wb["users"]
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            users.append({"uid": row[0], "name": row[1], "count": row[2]})
    return users


def user_buys_coffee(uid: str) -> dict | None:
    """
    Update a user's coffee count.
    Args:
        uid: The user ID.

    Returns:
        Updated user data or None if user not found.
    """
    wb = _get_spreadsheet()
    ws_users = wb["users"]
    ws_actions = wb["actions"]

    for row in ws_users.iter_rows(min_row=2):
        if row[0].value == uid:
            current_count = row[2].value or 0
            row[2].value = current_count + 1
            ws_actions.append([uid, row[1].value, datetime.now().isoformat()])
            wb.save(EXCEL_FILE)
            return {"uid": uid, "count": row[2].value}

    return None